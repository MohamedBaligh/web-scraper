from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42
ws.title = 'abc'
# Rows can also be appended
ws.append([11, 21, 31])
ws.append([21, 21, 31])
ws.append([31, 41, 51])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

ws = wb.create_sheet(title="Pi") 
ws.append([33,44,55])
ws.append([77,11,66])

# Save the file
wb.save("sample.xlsx")
